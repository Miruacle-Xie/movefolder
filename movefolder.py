# -*- coding: utf-8 -*-
"""
Created on Tue Oct 05 10:04:02 2022

@author: Administrator
"""
import shutil
import os
from openpyxl import load_workbook
from openpyxl import Workbook
# import pandas as pd


def main():
    folderpath = input("筛选文件夹:\n")
    folderpath = folderpath.replace("\'", "").replace("\"", "")
    filterfile = input("过滤的xlsx文件:\n")
    filterfile = filterfile.replace("\'", "").replace("\"", "")
    # df = pd.read_excel(filterfile, sheet_name=0)
    # df = df.iloc[:, 0]
    # print(df)
    wb = load_workbook(filterfile)
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]  # index为0为第一张表
    parentpath = folderpath + "\\"
    filterfolder = parentpath + "过滤文件夹"
    dontfind = ["文件夹中未找到的主题"]

    if not os.path.isdir(filterfolder):
        os.mkdir(filterfolder)
        # for i in range(len(df)):
    for i in range(1, ws.max_row + 1):
        try:
            # if os.path.isdir(parentpath + str(df[i])):
            if os.path.isdir(parentpath + str(ws.cell(i, 1).value)):
                # a = shutil.move(parentpath + str(df[i]), filterfolder)
                shutil.move(parentpath + str(ws.cell(i, 1).value), filterfolder)
                print(str(ws.cell(i, 1).value) + ":成功")
            else:
                 print(str(ws.cell(i, 1).value) + ":在文件夹中未找到({})".format(i))
                 dontfind.append(str(ws.cell(i, 1).value) + ":({})".format(i))
        except Exception as e:
            # print(parentpath + str(df[i]) + ":移动异常")
            print(str(ws.cell(i, 1).value) + ":移动异常")
            dontfind.append(str(ws.cell(i, 1).value) + ":移动异常({})".format(i))
            print(repr(e))
    if len(dontfind) > 0:
        resultfile = filterfolder + "\\moveresult.xlsx"
        if not os.path.exists(resultfile):
            wb = Workbook()
        else:
            os.remove(resultfile)
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]  # index为0为第一张表
        # sheet.append("文件夹中未找到的主题")
        # sheet.append(dontfind)
        for i in range(1, len(dontfind)+1):
            try:
                print(dontfind[i-1])
                ws.cell(i, 1).value = dontfind[i-1]
            except Exception as e:
                print(dontfind[i-1] + "写入异常")
                print(repr(e))
        wb.save(resultfile)
    input("过滤完成, 按回车结束")


def main1():
    folderpath = input("筛选文件夹:\n")
    folderpath = folderpath.replace("\'", "").replace("\"", "")
    filterfile = input("过滤的xlsx文件:\n")
    filterfile = filterfile.replace("\'", "").replace("\"", "")
    # df = pd.read_excel(filterfile, sheet_name=0)
    # df = df.iloc[:, 0]
    # print(df)
    wb = load_workbook(filterfile)
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]  # index为0为第一张表
    parentpath = folderpath + "\\"
    dontfind = ["文件夹中未找到的主题"]

    priority = [x.value for x in tuple(ws.columns)[1]]
    # print(priority)
    priority = list(set(priority))
    # print(priority)
    tmp = ""
    for a in range(len(priority)):
        tmp = tmp + str(a) + ":" + priority[a] + "  "
    print(tmp)
    selepriority = input("输入数字选择：\n".format(len(priority)+1))
    filterfolder = parentpath + priority[int(selepriority)]

    if not os.path.isdir(filterfolder):
        os.mkdir(filterfolder)
        # for i in range(len(df)):
    for i in range(1, ws.max_row + 1):
        try:
            # if os.path.isdir(parentpath + str(df[i])):
            if str(ws.cell(i, 2).value) in priority[int(selepriority)]:
                if os.path.isdir(parentpath + str(ws.cell(i, 1).value)):
                    # a = shutil.move(parentpath + str(df[i]), filterfolder)
                    shutil.move(parentpath + str(ws.cell(i, 1).value), filterfolder)
                    print(str(ws.cell(i, 1).value) + ":成功")
                else:
                     print(str(ws.cell(i, 1).value) + ":在文件夹中未找到({})".format(i))
                     dontfind.append(str(ws.cell(i, 1).value) + ":({})".format(i))
        except Exception as e:
            # print(parentpath + str(df[i]) + ":移动异常")
            print(str(ws.cell(i, 1).value) + ":移动异常")
            dontfind.append(str(ws.cell(i, 1).value) + ":移动异常({})".format(i))
            print(repr(e))
    if len(dontfind) > 0:
        resultfile = filterfolder + "\\moveresult.xlsx"
        if not os.path.exists(resultfile):
            wb = Workbook()
        else:
            os.remove(resultfile)
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]  # index为0为第一张表
        # sheet.append("文件夹中未找到的主题")
        # sheet.append(dontfind)
        for i in range(1, len(dontfind)+1):
            try:
                print(dontfind[i-1])
                ws.cell(i, 1).value = dontfind[i-1]
            except Exception as e:
                print(dontfind[i-1] + "写入异常")
                print(repr(e))
        wb.save(resultfile)
    input("过滤完成, 按回车结束")


if __name__ == "__main__":
    mode = input("请选择模式：\n1：移动文件夹  2:移动文件夹且按优先级\n")
    if int(mode) == 1:
        main()
    elif int(mode) == 2:
        main1()
    else:
        input("输入错误,请重新选择,按回车结束")
    # filterfile = input("过滤的xlsx文件:\n")
    # filterfile = filterfile.replace("\'", "").replace("\"", "")
    # # df = pd.read_excel(filterfile, sheet_name=0)
    # # df = df.iloc[:, 0]
    # # print(df)
    # wb = load_workbook(filterfile)
    # sheetnames = wb.sheetnames
    # ws = wb[sheetnames[0]]  # index为0为第一张表
    # print(type(ws.columns))
    # print(tuple(ws.columns)[1])
    # a = [x.value for x in tuple(ws.columns)[1]]
    # print(a)
    # b = list(set(a))
    # print(b)
    # tmp = ""
    # for i in range(len(b)):
    #     tmp = tmp + str(i) + ":" + b[i] + "  "
    # print(tmp)
